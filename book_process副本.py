import openpyxl

class book(object):
	def __init__(self,file):
		self.__file=file
	def read(self):
		self.__book=openpyxl.load_workbook(self.__file)
		self.__sheet=self.__book["Sheet1"]
		rowidx=1
		for row in self.__sheet.iter_rows(min_col=3,max_col=3,values_only=True):
			for count in row:
				if(count>0):
					print("init error")
					self.__book=None
					return None
		return self.__book
	def write(self,row,column,value):
		self.__sheet.cell(row,column,value)

	def record(self,strings,image_file):
		words={}
		rowidx=1
		for row in self.__sheet.iter_rows(max_col=1,values_only=True):
			for word in row:
				words[word]=rowidx
				# print(words[word],word)
			rowidx+=1
		for word in strings:
			if(word in words):
				rowidx=words[word]
				print("found %s at %d" % (word[:15],rowidx))
				resultcell=self.__sheet.cell(rowidx,2,"found")
				countcell=self.__sheet.cell(rowidx,3,self.__sheet.cell(rowidx,3).value+1)
				pathcell=self.__sheet.cell(rowidx,4,str(self.__sheet.cell(rowidx,4).value)+"\r"+image_file)

	def save(self):
		self.__book.save(self.__file)